#!/usr/bin/env python3
"""
ifc_info_excel.py – Analyse un ou **plusieurs** fichiers IFC (IFC2X3 / IFC4) et
produit des classeurs Excel structurés (Résumé, Niveaux, Arborescence).

* Fonctionne sur un **fichier** ou un **dossier** (option récursive).
* Coordonnées globales, latitude/longitude, NGF, tableaux formatés.
* Avril 2025 : prise en charge du mode dossier, option --merge, correctifs
  (conversion DMS, indentation, etc.).

Usage
-----
    # Fichier unique
    python ifc_info_excel.py modele.ifc [-o modele.xlsx]

    # Dossier complet → un xlsx par IFC
    python ifc_info_excel.py C:/BIM/IFC

    # Récursif + dossier de sortie dédié
    python ifc_info_excel.py C:/BIM/IFC -r -o C:/BIM/rapports

    # Fusionner dans un seul classeur
    python ifc_info_excel.py C:/BIM/IFC --merge -o rapport_global.xlsx

Dépendances
-----------
    pip install ifcopenshell pandas openpyxl>=3.1
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import ifcopenshell
except ImportError:
    sys.stderr.write("✖️  ifcopenshell introuvable – pip install ifcopenshell\n")
    raise SystemExit(1)

# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #

def human_readable_size(bytes_: int) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if bytes_ < 1024:
            return f"{bytes_:3.1f} {unit}"
        bytes_ /= 1024
    return f"{bytes_:3.1f} PB"


def get_length_factor(model) -> float:
    """Retourne 0.001 si le modèle est en millimètres, sinon 1.0."""
    try:
        ua = model.by_type("IfcUnitAssignment")[0].Units  # type: ignore[index]
        lu = next(u for u in ua if u.is_a("IfcSIUnit") and u.UnitType == "LENGTHUNIT")
        if getattr(lu, "Prefix", None) == "MILLI":
            return 0.001
    except Exception:
        pass
    return 1.0

# --------------------------------------------------------------------------- #
# Extraction IFC                                                              #
# --------------------------------------------------------------------------- #

def get_levels(model):
    levels = [(s.Name, getattr(s, "Elevation", None)) for s in model.by_type("IfcBuildingStorey")]
    return sorted(levels, key=lambda t: (t[1] or 0))


def dms_to_dd(dms):
    """Convertit un tuple/list DMS → degrés décimaux.

    Accepte 3 ou 4 éléments : [deg, min, sec] ou [deg, min, sec, 1e‑6 sec].
    Retourne None si dms est None.
    """
    if dms is None:
        return None
    parts = list(dms) + [0] * (4 - len(dms))  # pad à 4 éléments
    deg, minute, sec, micro = parts[:4]
    sign = -1 if deg < 0 else 1
    return sign * (abs(deg) + minute / 60 + (sec + micro / 1e6) / 3600)


def get_site_geolocation(model):
    site = next(iter(model.by_type("IfcSite")), None)
    if not site:
        return {}
    return {
        "latitude": dms_to_dd(site.RefLatitude) if site.RefLatitude else None,
        "longitude": dms_to_dd(site.RefLongitude) if site.RefLongitude else None,
        "elevation_m": site.RefElevation,
    }


def get_global_coords(model):
    factor = get_length_factor(model)
    try:
        map_conv = next(iter(model.by_type("IfcMapConversion")))
    except (RuntimeError, StopIteration):
        map_conv = None
    if map_conv:
        return {
            "X_global_m": map_conv.Eastings * factor,
            "Y_global_m": map_conv.Northings * factor,
            "Z_global_m": (map_conv.OrthogonalHeight or 0) * factor if map_conv.OrthogonalHeight is not None else None,
        }
    site = next(iter(model.by_type("IfcSite")), None)
    if site and site.ObjectPlacement and site.ObjectPlacement.is_a("IfcLocalPlacement"):
        loc = site.ObjectPlacement.RelativePlacement.Location  # type: ignore[attr-defined]
        x, y, *rest = list(loc.Coordinates) + [None, None]
        z = rest[0]
        return {
            "X_global_m": x * factor if x is not None else None,
            "Y_global_m": y * factor if y is not None else None,
            "Z_global_m": z * factor if z is not None else None,
        }
    return {}


def flatten_hierarchy(model):
    rels = {rel.RelatingObject.GlobalId: rel.RelatedObjects for rel in model.by_type("IfcRelAggregates")}
    project = next(iter(model.by_type("IfcProject")), None)
    rows: List[Dict[str, str | int]] = []

    def walk(ent, path):
        name = ent.Name or ent.LongName or ent.GlobalId
        new_path = path + [f"{ent.is_a()}:{name}"]
        rows.append({"Type": ent.is_a(), "Nom": name, "Profondeur": len(path), "Chemin": "/".join(new_path)})
        for child in rels.get(ent.GlobalId, []):
            walk(child, new_path)

    if project:
        walk(project, [])
    return rows

# --------------------------------------------------------------------------- #
# Excel helpers                                                               #
# --------------------------------------------------------------------------- #

def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int):
    for c, col in enumerate(df.columns):
        cell = ws.cell(row=start_row + 1, column=start_col + c + 1, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
    for r_idx, row in enumerate(df.itertuples(index=False, name=None)):
        for c_idx, value in enumerate(row):
            ws.cell(row=start_row + 2 + r_idx, column=start_col + c_idx + 1, value=value)


def add_table_and_resize(ws, df: pd.DataFrame, table_name: str, start_row: int, start_col: int):
    n_rows, n_cols = df.shape
    first_row = start_row + 1
    last_row = first_row + n_rows
    first_col = start_col + 1
    last_col = first_col + n_cols - 1
    ref = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    for idx, col in enumerate(df.columns):
        col_letter = get_column_letter(first_col + idx)
        max_len = max([len(str(col))] + df[col].astype(str).str.len().tolist())
        ws.column_dimensions[col_letter].width = max_len + 2

# --------------------------------------------------------------------------- #
# Génération d'un classeur pour UN IFC                                        #
# --------------------------------------------------------------------------- #

def build_workbook(ifc_path: Path) -> Workbook:
    """Construit un classeur Excel pour un IFC unique, avec 4 onglets :
    Résumé, Niveaux, Arborescence et Entités.
    """
    model = ifcopenshell.open(ifc_path.as_posix())
    factor = get_length_factor(model)

    # -------- Résumé ------------------------------------------------------- #
    geo = get_site_geolocation(model)
    glob = get_global_coords(model)
    summary_pairs = [
        ("fichier", ifc_path.name),
        ("taille", human_readable_size(ifc_path.stat().st_size)),
        ("schéma", model.schema if isinstance(model.schema, str) else model.schema.identifier),
    ] + list(glob.items()) + list(geo.items())
    summary_df = pd.DataFrame(summary_pairs, columns=["Propriété", "Valeur"])

    # -------- Niveaux ------------------------------------------------------ #
    site_z = glob.get("Z_global_m")
    lvls = get_levels(model)
    level_alt = [(e * factor if e is not None else None) for _, e in lvls]
    level_ngf = [(a + site_z) if (a is not None and site_z is not None) else None for a in level_alt]
    levels_df = pd.DataFrame({
        "Nom": [n for n, _ in lvls],
        "Altimétrie locale (m)": level_alt,
        "Altimétrie NGF (m)": level_ngf,
    })

    # -------- Arborescence ------------------------------------------------- #
    hierarchy_df = pd.DataFrame(flatten_hierarchy(model))

    # -------- Entités ------------------------------------------------------ #
    entity_counter: Dict[tuple[str, str | None], int] = {}
    for obj in model.by_type("IfcProduct"):
        ent = obj.is_a()
        typ = None
        # Préférence : PredefinedType puis ObjectType
        if hasattr(obj, "PredefinedType") and obj.PredefinedType is not None:
            typ = str(obj.PredefinedType)
        elif getattr(obj, "ObjectType", None):
            typ = obj.ObjectType
        key = (ent, typ)
        entity_counter[key] = entity_counter.get(key, 0) + 1

    ent_rows = [
        {"Entité IFC": k[0], "Type": k[1] or "—", "Nombre": v}
        for k, v in sorted(entity_counter.items(), key=lambda x: (-x[1], x[0][0]))
    ]
    entities_df = pd.DataFrame(ent_rows)

    # -------- Workbook & sheets ------------------------------------------- #
    wb = Workbook()
    wb.remove(wb.active)
    sheets_data = [
        ("Résumé", summary_df),
        ("Niveaux", levels_df),
        ("Arborescence", hierarchy_df),
        ("Entités", entities_df),
    ]
    for sheet, df in sheets_data:
        ws = wb.create_sheet(sheet)
        write_df(ws, df, 1, 1)
        add_table_and_resize(ws, df, f"{sheet}Tbl", 1, 1)

        if sheet == "Entités":
            # surlignage des IfcBuildingElementProxy
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            orange_font = Font(color="C55A11", italic=True)
            start_col_idx = 2  # B
            attention_col_idx = start_col_idx + len(df.columns)  # juste après la dernière colonne
            for row_offset, ent_val in enumerate(df["Entité IFC"]):
                if ent_val == "IfcBuildingElementProxy":
                    excel_row = 3 + row_offset  # 1‑based : ligne 2 = en‑tête, données démarrent ligne 3
                    # Coloriage de la ligne
                    for col_off in range(len(df.columns)):
                        ws.cell(row=excel_row, column=start_col_idx + col_off).fill = orange_fill
                    # Message « Attention »
                    att_cell = ws.cell(row=excel_row, column=attention_col_idx)
                    att_cell.value = "Attention"
                    att_cell.font = orange_font
    return wb

# --------------------------------------------------------------------------- #
# Batch helpers                                                               #
# --------------------------------------------------------------------------- #

def save_wb(wb: Workbook, path: Path):
    """Save workbook to path, creating parent directory if needed."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  ✔ {path.name}")


def discover_ifcs(folder: Path, recursive: bool) -> List[Path]:
    """Return .ifc files in folder (optionally recursive)."""
    pattern = "**/*.ifc" if recursive else "*.ifc"
    return sorted(p for p in folder.glob(pattern) if p.is_file())


def process_single(ifc_path: Path, out_dir: Path | None, merge: bool, merge_wb: Workbook | None):
    """Process one IFC: create Excel either standalone or merged."""
    wb = build_workbook(ifc_path)

    if merge and merge_wb is not None:
        # copy sheets into global workbook
        for sheet in wb.sheetnames:
            src = wb[sheet]
            dst = merge_wb.create_sheet(f"{sheet}_{ifc_path.stem}")
            for row in src.rows:
                for cell in row:
                    dst[cell.coordinate].value = cell.value
                    dst[cell.coordinate]._style = cell._style
        return

    target = (out_dir or ifc_path.parent) / ifc_path.with_suffix(".xlsx").name
    save_wb(wb, target)

# --------------------------------------------------------------------------- #
# CLI & batch processing                                                      #
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(description="Export infos IFC → Excel (fichier ou dossier)")
    parser.add_argument("src", type=Path, help="Fichier IFC ou dossier")
    parser.add_argument("-o", "--output", type=Path, help="Dossier ou fichier de sortie (voir --merge)")
    parser.add_argument("-r", "--recursive", action="store_true", help="Descend dans les sous-dossiers")
    parser.add_argument("--merge", action="store_true", help="Combine tous les résultats dans un seul .xlsx")
    args = parser.parse_args()

    if args.src.is_file():
        wb = build_workbook(args.src)
        target = args.output or args.src.with_suffix(".xlsx")
        save_wb(wb, target)
        return

    if not args.src.is_dir():
        sys.exit("❌  Chemin source introuvable")

    ifc_files = discover_ifcs(args.src, args.recursive)
    if not ifc_files:
        sys.exit("❌  Aucun .ifc trouvé dans ce dossier")

    print(f"▶ Analyse de {len(ifc_files)} IFC …")

    merge_wb = Workbook() if args.merge else None
    if merge_wb and merge_wb.active:
        merge_wb.remove(merge_wb.active)

    out_dir = None if args.merge else (args.output if args.output and args.output.suffix == "" else args.output)

    for f in ifc_files:
        print(f"- {f.name}")
        process_single(f, out_dir, args.merge, merge_wb)

    if args.merge and merge_wb:
        out_path = args.output or (args.src / "rapport_ifc.xlsx")
        save_wb(merge_wb, out_path)

    print("✅  Terminé.")


if __name__ == "__main__":
    main()
